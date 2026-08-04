#!/usr/bin/env python3
"""Build brooks-barcode-feed.json from the Brooks UPC workbook.

Brooks is the one brand with no barcode source wired into the tool. ASICS and
ON have theirs baked into barcode-data.js, Hoka carries barcodes in its own
feed, and Saucony was backfilled from the Catalog UPCs export. Brooks products
created by the tool therefore go to Shopify bare, which means a shoe that
cannot be scanned at the till.

The source is the Brooks UPC workbook, "S2027 UPC Codes 051826.xlsx", one sheet
per season back to S2023. It IS a supplier file and it DOES carry wholesale
prices, so it stays in the gitignored set and never gets committed. What this
writes out is barcodes only:

    { "<ITEM_NUMBER>": "<UPC>", ... }        eg "1104962E020.070": "195394885736"

ITEM_NUMBER is Brooks's own variant code and decomposes as

    110464  2E   020    .   070
    style   dim  color      size, in tenths

which is exactly the SKU shape on the older Brooks products in the store. The
tool-created ones use a different dialect ("110496015-020-750-2E"), so the
consumer converts before looking up; see brooksItemNumber() in
tools/backfill-brooks-barcodes.mjs.

Run:  python3 tools/build-brooks-barcode-feed.py
      python3 tools/build-brooks-barcode-feed.py --src "path/to/UPC Codes.xlsx"

House style: no em dashes. Use commas, periods, or the word "to".
"""

import glob
import json
import os
import sys
import warnings

warnings.filterwarnings("ignore")

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required:  python3 -m pip install openpyxl")

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, "brooks-barcode-feed.json")

# Only these columns are read. WHOLESALE, RETAIL_PRICE and the Canada prices are
# deliberately never touched, so nothing priced can leak into the output.
WANT = ("ITEM_NUMBER", "UPC")


def find_source(argv):
    i = argv.index("--src") if "--src" in argv else -1
    if i >= 0:
        return argv[i + 1]
    # The workbook lives in the repo root or in barcodes/, and its name carries
    # the season it was pulled for, so match on the stable part.
    for pattern in ("*UPC Codes*.xlsx", os.path.join("barcodes", "*UPC Codes*.xlsx")):
        hits = sorted(glob.glob(os.path.join(ROOT, pattern)))
        if hits:
            return hits[-1]
    sys.exit(
        "No Brooks UPC workbook found. Put 'S2027 UPC Codes ....xlsx' in the repo\n"
        "root or in barcodes/, or pass --src <path>."
    )


def main(argv):
    src = find_source(argv)
    print(f"Reading {os.path.basename(src)}")
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)

    out = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        header = None
        kept = 0
        for row in ws.iter_rows(values_only=True):
            if header is None:
                header = {str(v).strip(): i for i, v in enumerate(row) if v is not None}
                if not all(c in header for c in WANT):
                    # Not a UPC sheet, skip it rather than guess at its columns.
                    header = None
                    break
                continue
            item = row[header["ITEM_NUMBER"]] if header["ITEM_NUMBER"] < len(row) else None
            code = row[header["UPC"]] if header["UPC"] < len(row) else None
            if not item or not code:
                continue
            item = str(item).strip()
            code = str(code).strip()
            if not item or not code:
                continue
            # First season wins. The sheets run newest to oldest and a repeated
            # ITEM_NUMBER is the same physical shoe, so this is stable either way.
            if item not in out:
                out[item] = code
                kept += 1
        print(f"  {sheet}: {kept} barcodes")
    wb.close()

    with open(OUT, "w") as f:
        json.dump(out, f, separators=(",", ":"), sort_keys=True)
    size_mb = os.path.getsize(OUT) / 1048576
    print(f"\nWrote {len(out)} barcodes to {os.path.relpath(OUT, ROOT)} ({size_mb:.1f} MB)")
    print("Barcodes only, no pricing. The file is gitignored; regenerate it when")
    print("a new season's workbook arrives.")


if __name__ == "__main__":
    main(sys.argv[1:])
