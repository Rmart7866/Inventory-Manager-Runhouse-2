#!/usr/bin/env python3
# Build barcode-data.js from the raw supplier files in ../barcodes/.
#
# The raw files (ASICS order-form CSVs, the ON pricat xlsx) are gitignored and
# NEVER shipped: they contain wholesale prices and order data. This script
# extracts ONLY the barcode map (style/color/size -> EAN) into barcode-data.js,
# which the app loads so users never re-upload barcodes.
#
# Refresh after dropping new files in barcodes/:  python3 tools/build-barcodes.py

import csv, re, os, glob, json
HERE = os.path.dirname(os.path.abspath(__file__))
BARCODES = os.path.join(HERE, '..', 'barcodes')
OUT = os.path.join(HERE, '..', 'barcode-data.js')

def nsize(s):
    """Normalize a size the SAME way the converters' _normUpcSize does in JS:
    '8'->'8', '8.5'->'8.5', 'K12.5'->'K12.5'. Returns None if not a shoe size."""
    s = str(s).strip()
    m = re.match(r'^(K?)(\d{1,2}(\.5)?)$', s, re.I)
    if not m:
        return None
    n = float(m.group(2))
    num = str(int(n)) if n == int(n) else str(n)
    return ('K' if m.group(1) else '') + num

def nsize_apparel(s):
    """Normalize an APPAREL size the same way OnApparelConverter.normalizeSize
    does in JS. The two must agree exactly or a barcode never finds its variant.

    The store, the pricat and ordinary typing disagree about how to write one
    size: 'X-Small' / 'XS' / 'XSmall', and at the top end '2X-Large' / 'XXL' /
    '2XL'. Collapse them. Bra cup ranges ('S D-DD') are kept as a suffix, or
    every cup of a size would share one key. Returns None if it is not a size.
    """
    s = str(s).strip().upper()
    if not s:
        return None
    if s in ('-', 'OS') or re.match(r'^ONE\s*SIZE$', s):
        return 'OS'                       # the pricat writes one-size as "-"
    cup = ''
    m = re.search(r'\(?\s*([A-Z])\s*-\s*(DD|[A-Z])\s*\)?\s*$', s)
    if m and m.group(1) in 'ABCD':
        cup = '|' + m.group(1) + '-' + m.group(2)
        s = s[:m.start()].strip()
    s = re.sub(r'[\s()-]', '', s)
    m = re.match(r'^(\d+)?(X*)(SMALL|MEDIUM|LARGE|S|M|L)$', s)
    if not m:
        return None
    word = m.group(3)
    base = 'S' if word in ('SMALL', 'S') else ('M' if word in ('MEDIUM', 'M') else 'L')
    if base == 'M':
        return 'M' + cup
    mult = int(m.group(1)) if m.group(1) else len(m.group(2) or '')
    if not mult:
        return base + cup
    return ('X' if mult == 1 else str(mult) + 'X') + base + cup

def nprice(s):
    """MSRP only (retail, public). '$230' -> '230.00'. Ignores blanks."""
    m = re.search(r'(\d+(?:\.\d+)?)', str(s))
    if not m:
        return None
    return '%.2f' % float(m.group(1))

# ---- ASICS: order-form CSVs. barcode key "TradingCode-ColorCode|USsize" -> EAN.
# price key "TradingCode-ColorCode" -> Suggested Retail Price (MSRP, public). ----
asics = {}
asics_price = {}
for f in glob.glob(os.path.join(BARCODES, '*.csv')):
    try:
        rows = list(csv.reader(open(f, newline='', encoding='utf-8-sig')))
    except Exception:
        continue
    hi = next((i for i, r in enumerate(rows[:8]) if any('EAN' in str(c) for c in r)), None)
    if hi is None:
        continue
    idx = {str(n).strip(): j for j, n in enumerate(rows[hi])}
    if 'EAN code' not in idx or 'Trading code' not in idx:
        continue
    for r in rows[hi + 1:]:
        def g(n):
            j = idx.get(n)
            return '' if j is None or j >= len(r) else str(r[j]).strip()
        ean, tc, cc, sz = g('EAN code'), g('Trading code').upper(), g('Color code'), nsize(g('Size US'))
        if ean and tc and cc and sz:
            asics[tc + '-' + cc + '|' + sz] = ean
        if tc and cc:
            p = nprice(g('Suggested Retail Price'))
            if p:
                asics_price[tc + '-' + cc] = p

# ---- ON: pricat xlsx. barcode key "ItemCode|USsize" -> EAN.
# price key "ItemCode" -> Retail Price (MSRP, public). ----
on = {}
on_apparel = {}
on_price = {}
try:
    from openpyxl import load_workbook
    for f in glob.glob(os.path.join(BARCODES, '*.xlsx')):
        ws = load_workbook(f, read_only=True, data_only=True).worksheets[0]
        it = ws.iter_rows(values_only=True)
        hdr = list(next(it)); idx = {n: i for i, n in enumerate(hdr)}
        if 'Item Code' not in idx or 'EAN Barcode' not in idx:
            continue
        for r in it:
            def g(n):
                i = idx.get(n)
                return '' if i is None or i >= len(r) or r[i] is None else str(r[i]).strip()
            code, ean = g('Item Code').upper(), g('EAN Barcode')
            raw = g('US Size') or g('Size')
            sz = nsize(raw)
            if code and ean and sz:
                on[code + '|' + sz] = ean
            # APPAREL. The same pricat carries the garments, 2,792 rows of them,
            # but their sizes are words (XS, S, M, XXL, "S D-DD") so nsize above
            # rejects every one and they were all being discarded. Footwear codes
            # start with 3 and apparel with 1, so they cannot collide; keep them
            # in their own map, keyed the way OnApparelConverter normalizes.
            elif code.startswith('1') and ean:
                asz = nsize_apparel(raw)
                if asz:
                    on_apparel[code + '|' + asz] = ean
            if code:
                p = nprice(g('Retail Price'))
                if p:
                    on_price[code] = p
except ImportError:
    print('  (openpyxl not installed — skipping ON pricat)')

# ---- BROOKS: the UPC workbook, one sheet per season. barcode key is Brooks's
# own ITEM_NUMBER ("1104962E020.070") -> UPC. The converter decodes its scraper
# SKU into that shape before looking up, see brooksItemNumber in
# brooks-converter.js and in tools/backfill-brooks-barcodes.mjs.
#
# ONLY THE NEWEST SEASONS. The workbook goes back to S2023 and holds 52,936
# rows, which is 1.6 MB of JS the browser would load on every page view. What
# the tool actually needs is the seasons it can still create products from, so
# the older sheets are dropped. Widen BROOKS_SEASONS if a backfill needs them,
# or use tools/build-brooks-barcode-feed.py, which keeps the lot out of band.
BROOKS_SEASONS = ('S2027', 'F2026', 'S2026', 'F2025')
brooks = {}
try:
    from openpyxl import load_workbook
    # The workbook lives in barcodes/ like the others, or in the repo root where
    # it was first dropped.
    books = glob.glob(os.path.join(BARCODES, '*UPC Codes*.xlsx')) \
        + glob.glob(os.path.join(HERE, '..', '*UPC Codes*.xlsx'))
    for f in sorted(books):
        wb = load_workbook(f, read_only=True, data_only=True)
        for sheet in wb.sheetnames:
            if sheet not in BROOKS_SEASONS:
                continue
            it = wb[sheet].iter_rows(values_only=True)
            hdr = list(next(it))
            idx = {str(n).strip(): i for i, n in enumerate(hdr) if n is not None}
            if 'ITEM_NUMBER' not in idx or 'UPC' not in idx:
                continue
            for r in it:
                def g(n):
                    i = idx.get(n)
                    return '' if i is None or i >= len(r) or r[i] is None else str(r[i]).strip()
                item, upc = g('ITEM_NUMBER').upper(), g('UPC')
                # First season wins. Sheets run newest first and a repeated
                # ITEM_NUMBER is the same physical shoe.
                if item and upc and item not in brooks:
                    brooks[item] = upc
        wb.close()
except ImportError:
    print('  (openpyxl not installed — skipping Brooks UPC workbook)')

data = {'asics': asics, 'on': on, 'onApparel': on_apparel, 'brooks': brooks,
        'prices': {'asics': asics_price, 'on': on_price}}
body = (
    '// AUTO-GENERATED by tools/build-barcodes.py — do not edit by hand.\n'
    '// Barcode maps only (style/color/size -> EAN); no pricing. The raw supplier\n'
    '// files stay in the gitignored barcodes/ folder. Regenerate after adding\n'
    '// files:  python3 tools/build-barcodes.py\n'
    'var BarcodeData = ' + json.dumps(data, separators=(',', ':')) + ';\n'
    "if (typeof module !== 'undefined' && module.exports) module.exports = BarcodeData;\n"
)
open(OUT, 'w').write(body)
print('barcodes  asics: %d  on: %d  onApparel: %d  brooks: %d   |   prices  asics: %d  on: %d   ->  barcode-data.js (%.1f KB)' % (
    len(asics), len(on), len(on_apparel), len(brooks), len(asics_price), len(on_price), os.path.getsize(OUT) / 1024))
